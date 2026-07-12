#!/usr/bin/env python3
"""
cdgc_create_dq_occurrences.py

Three-phase DQ occurrence pipeline — generate, import, and link in one run:

  Phase 1  Read 13_DQ_Rule_Template_PATCHED.xlsx, fetch all scanned columns,
           build Primary Data Element paths from core.location
  Phase 2  Match rules to columns, write 15_DQ_Rule_Occurrence.xlsx, preview → CONFIRM
  Phase 3  Import File 15 via API (poll to COMPLETED), then PATCH all
           template→occurrence relationships

Usage:
  python3 cdgc_create_dq_occurrences.py              # Operation=Create (first run)
  python3 cdgc_create_dq_occurrences.py --update     # Operation=Update (re-run)

To re-run after changes: delete existing occurrences first with
  python3 cdgc_delete_dq_occurrences.py
then re-run this script.
"""
import argparse
import getpass
import json
import sys
import time
from pathlib import Path

import requests

parser = argparse.ArgumentParser(add_help=False)
parser.add_argument("--update", action="store_true", help="Set Operation=Update instead of Create")
args_parsed, _ = parser.parse_known_args()
OPERATION = "Update" if args_parsed.update else "Create"

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("pip3 install openpyxl")

LOGIN_URL    = "https://dmp-us.informaticacloud.com"
ORG_URL      = "https://idmc-api.dmp-us.informaticacloud.com"
COLUMN_CLASS = "com.infa.odin.models.relational.Column"
ASSOC_TYPE   = "com.infa.ccgf.models.governance.relatedRuleTemplateRuleInstance"

# ── Inputs ────────────────────────────────────────────────────────────────────
import_dir_raw = input("Import directory (e.g. ~/Downloads/CDGC_Import_MyClient): ").strip()
IMPORT_DIR = Path(import_dir_raw).expanduser()
OCC_PREFIX = input("Occurrence reference ID prefix (e.g. DQOCC): ").strip()

TEMPLATE_FILE   = IMPORT_DIR / "13_DQ_Rule_Template_PATCHED.xlsx"
OCCURRENCE_FILE = IMPORT_DIR / (
    "15_DQ_Rule_Occurrence_UPDATE.xlsx" if OPERATION == "Update"
    else "15_DQ_Rule_Occurrence.xlsx"
)

METHOD_MAP = {
    "technicalscript":               "TechnicalScript",
    "technical script":              "TechnicalScript",
    "businessextract":               "BusinessExtract",
    "business extract":              "BusinessExtract",
    "systemfunction":                "SystemFunction",
    "system function":               "SystemFunction",
    "informaticaclouddataquality":   "InformaticaCloudDataQuality",
    "informatica cloud data quality":"InformaticaCloudDataQuality",
}

# ── Auth ──────────────────────────────────────────────────────────────────────
username = input("IDMC Username: ").strip()
password = getpass.getpass("IDMC Password: ")

r = requests.post(f"{LOGIN_URL}/identity-service/api/v1/Login",
    json={"username": username, "password": password}, timeout=30)
r.raise_for_status()
sid, oid = r.json()["sessionId"], r.json()["orgId"]
r2 = requests.get(
    f"{LOGIN_URL}/identity-service/api/v1/jwt/Token?client_id=idmc_api&nonce=1234",
    headers={"IDS-SESSION-ID": sid}, cookies={"USER_SESSION": sid}, timeout=30)
jwt = r2.json().get("token") or r2.json().get("jwt_token")
H_base = {"Authorization": f"Bearer {jwt}", "X-INFA-ORG-ID": oid}
H      = {**H_base, "Content-Type": "application/json"}
print("✓ Authenticated\n")

# ══ PHASE 1: Read templates, fetch columns, build PDE paths ══════════════════

print("Phase 1 — Reading 13_DQ_Rule_Template_PATCHED.xlsx...")
wb_tmpl = load_workbook(TEMPLATE_FILE)
ws_tmpl = wb_tmpl.active
hdrs = [str(c.value or "").strip() for c in ws_tmpl[1]]
print(f"  Headers: {hdrs}\n")

rules = []
for row in ws_tmpl.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    d = dict(zip(hdrs, row))
    if str(d.get("Operation") or "").strip().lower() == "delete":
        continue
    input_port = str(d.get("Input Port Name") or "").strip()
    if not input_port:
        print(f"  SKIP (no Input Port Name): {d.get('Reference ID')} — {d.get('Name')}")
        continue
    raw_method = str(d.get("Measuring Method") or "TechnicalScript").strip()
    method     = METHOD_MAP.get(raw_method.lower(), raw_method)
    rules.append({
        "ref_id":       str(d.get("Reference ID") or "").strip(),
        "name":         str(d.get("Name") or "").strip(),
        "dimension":    str(d.get("Dimension") or "Completeness").strip(),
        "criticality":  str(d.get("Criticality") or "Medium").strip(),
        "method":       method,
        "tech_ref":     str(d.get("Technical Rule Reference") or "").strip(),
        "target":       d.get("Target") if d.get("Target") is not None else 95,
        "threshold":    d.get("Threshold") if d.get("Threshold") is not None else 90,
        "input_port":   input_port,
        "primary_glossary": str(d.get("Primary Glossary") or "").strip(),
    })

print(f"  {len(rules)} rules loaded:")
for r_ in rules:
    print(f"    [{r_['ref_id']}] {r_['name']}  →  column: {r_['input_port']}")

print("\nFetching all scanned columns from CDGC...")
all_cols = []
offset   = 0
while True:
    r3 = requests.post(
        f"{ORG_URL}/data360/search/v1/assets?knowledgeQuery=*&segments=summary",
        headers=H,
        json={"from": offset, "size": 100,
              "filterSpec": [{"type": "simple", "attribute": "core.classType",
                              "values": [COLUMN_CLASS]}]}, timeout=30)
    r3.raise_for_status()
    hits = r3.json().get("hits", [])
    for h in hits:
        s    = h.get("summary") or {}
        name = s.get("core.name", "")
        loc  = s.get("core.location", "")
        if name and loc:
            all_cols.append({"name": name, "loc": loc})
    if len(hits) < 100:
        break
    offset += 100
    time.sleep(0.3)

print(f"  {len(all_cols)} columns found")
print("\n  Sample core.location values (first 5):")
for c in all_cols[:5]:
    print(f"    name={c['name']!r}  loc={c['loc']!r}")


def get_catalog_source_name(uuid):
    """Resolve the human-readable catalog source name from its internal UUID."""
    try:
        r = requests.get(
            f"{ORG_URL}/data360/search/v1/assets/{uuid}?scheme=internal&segments=summary",
            headers=H, timeout=30)
        if r.status_code == 200:
            name = (r.json().get("summary") or {}).get("core.name", "")
            if name:
                return name
    except Exception:
        pass
    return uuid


def build_pde(loc, catalog_name):
    """
    core.location format:   UUID://UUID/DB/Schema/Table/Column
    Required import format: CatalogSourceName://DB/Schema/Table/Column
    """
    if "://" not in loc:
        return ""
    after_scheme = loc.split("://", 1)[1]
    slash_pos = after_scheme.find("/")
    if slash_pos == -1:
        return ""
    path = after_scheme[slash_pos + 1:]
    if not path:
        return ""
    return f"{catalog_name}://{path}"


catalog_name = ""
if all_cols:
    first_loc = all_cols[0]["loc"]
    if "://" in first_loc:
        catalog_uuid = first_loc.split("://", 1)[0]
        print(f"\nResolving catalog source name for UUID: {catalog_uuid} ...")
        resolved = get_catalog_source_name(catalog_uuid)
        if resolved and resolved != catalog_uuid:
            catalog_name = resolved
            print(f"  → Catalog source name: {catalog_name}")
        else:
            catalog_name = input("  Could not resolve. Enter catalog source name manually (MCC → Catalog Sources): ").strip()
            print(f"  → Using: {catalog_name}")

col_by_name = {}
for c in all_cols:
    key = c["name"].upper()
    pde = build_pde(c["loc"], catalog_name)
    if not pde:
        continue
    path_after_scheme = c["loc"].split("://", 1)[1] if "://" in c["loc"] else c["loc"]
    slash_pos = path_after_scheme.find("/")
    path_only = path_after_scheme[slash_pos + 1:] if slash_pos != -1 else path_after_scheme
    parts = [p for p in path_only.split("/") if p]
    table = parts[-2] if len(parts) >= 2 else (parts[-1] if parts else "UNKNOWN")
    entry = {"pde": pde, "table": table}
    if key not in col_by_name:
        col_by_name[key] = []
    if pde not in [e["pde"] for e in col_by_name[key]]:
        col_by_name[key].append(entry)

# ══ PHASE 2: Match rules to columns, build File 15, prompt CONFIRM ════════════

print("\nPhase 2 — Matching rules to columns...")
occurrences = []
occ_counter = 1
warnings    = []

for rule in rules:
    port_key = rule["input_port"].upper()
    matches  = col_by_name.get(port_key, [])
    if not matches:
        msg = f"  WARNING: No columns match Input Port Name '{rule['input_port']}' (rule {rule['ref_id']} — {rule['name']})"
        print(msg)
        warnings.append(msg)
        continue
    print(f"  {rule['ref_id']} ({rule['input_port']}): {len(matches)} column(s)")
    for m in matches:
        occ_name = f"{rule['name']} — {m['table']}.{rule['input_port']}"
        occ_ref  = f"{OCC_PREFIX}-{occ_counter}"
        occ_counter += 1
        occurrences.append({
            "ref_id":     occ_ref,
            "name":       occ_name,
            "dimension":  rule["dimension"],
            "criticality":rule["criticality"],
            "method":     rule["method"],
            "tech_ref":   rule["tech_ref"],
            "target":     rule["target"],
            "threshold":  rule["threshold"],
            "pde":        m["pde"],
            "table":      m["table"],
            "rule_name":  rule["name"],
            "input_port": rule["input_port"],
            "tpl_ref":    rule["ref_id"],
        })

print(f"\n  Occurrences to create: {len(occurrences)}")

print("\n  Preview:")
print(f"  {'Ref ID':<12} {'Table':<30} {'Column':<20} {'PDE (truncated)'}")
print(f"  {'-'*12} {'-'*30} {'-'*20} {'-'*40}")
for occ in occurrences[:20]:
    col = occ["pde"].split("/")[-1]
    pde_short = occ["pde"][:60] + "..." if len(occ["pde"]) > 60 else occ["pde"]
    print(f"  {occ['ref_id']:<12} {occ['table']:<30} {col:<20} {pde_short}")
if len(occurrences) > 20:
    print(f"  ... and {len(occurrences) - 20} more")

if warnings:
    print(f"\n  {len(warnings)} rules had no matching column (will be skipped in import):")
    for w in warnings:
        print(f"  {w}")

print(f"""
This script will now:
  1. Write   {OCCURRENCE_FILE.name}  ({len(occurrences)} rows)
  2. Import  via CDGC bulk import API (POST -> poll to COMPLETED)
  3. Link    {len(occurrences)} template->occurrence relationships via PATCH API
""")
confirm = input("Proceed? [yes/no]: ").strip().lower()
if confirm not in ("yes", "y"):
    print("Aborted.")
    sys.exit(0)

# ── Write 15_DQ_Rule_Occurrence.xlsx ─────────────────────────────────────────
print(f"\nWriting {OCCURRENCE_FILE.name}...")
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Data Quality Rule Occurrence"

HDR_COLS = [
    ("Reference ID",                          18),
    ("Name",                                  50),
    ("Description",                           20),
    ("Criticality",                           14),
    ("Dimension",                             16),
    ("Exception File Path",                   20),
    ("Frequency",                             14),
    ("Input Port Name",                       20),
    ("Lifecycle",                             14),
    ("Measuring Method",                      24),
    ("Output Port Name",                      20),
    ("Scanned Time",                          20),
    ("Score",                                 10),
    ("Technical Description",                 24),
    ("Technical Rule Reference",              28),
    ("Target",                                10),
    ("Threshold",                             12),
    ("Total Rows",                            14),
    ("Failed Rows",                           14),
    ("Primary Data Element",                  70),
    ("Secondary Data Element",                30),
    ("Operation",                             12),
    ("Stakeholder: Governance Owner",         30),
    ("Stakeholder: Governance Administrator", 30),
]

hdr_fill = PatternFill("solid", fgColor="1B2A4A")
hdr_font = Font(color="FFFFFF", bold=True, size=11)
row_font = Font(size=10)

for i, (label, width) in enumerate(HDR_COLS, 1):
    c = ws_out.cell(1, i, label)
    c.font      = hdr_font
    c.fill      = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_out.column_dimensions[get_column_letter(i)].width = width
ws_out.row_dimensions[1].height = 22

for occ in occurrences:
    row_data = [
        occ["ref_id"],
        occ["name"],
        "",
        occ["criticality"],
        occ["dimension"],
        "",
        "",
        occ["input_port"],
        "",
        occ["method"],
        "Output" if occ["method"] == "InformaticaCloudDataQuality" else "",
        "",
        "",
        "",
        occ["tech_ref"],
        occ["target"],
        occ["threshold"],
        "",
        "",
        occ["pde"],
        "",
        OPERATION,
        "",
        "",
    ]
    ws_out.append(row_data)
    rn = ws_out.max_row
    for ci in range(1, len(HDR_COLS) + 1):
        ws_out.cell(rn, ci).font      = row_font
        ws_out.cell(rn, ci).alignment = Alignment(vertical="center")
    ws_out.row_dimensions[rn].height = 16

wb_out.save(OCCURRENCE_FILE)
print(f"  Saved: {OCCURRENCE_FILE}")

# ══ PHASE 3a: Import File 15 via API ═════════════════════════════════════════

print(f"\nPhase 3a — Importing {OCCURRENCE_FILE.name}...")
with open(OCCURRENCE_FILE, "rb") as f:
    resp = requests.post(
        f"{ORG_URL}/data360/content/import/v1/assets",
        headers=H_base,
        files={
            "file": (OCCURRENCE_FILE.name, f,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            "config": (None, '{"validationPolicy":"CONTINUE_ON_ERROR_WARNING"}',
                       "application/json"),
        },
        timeout=60)

if resp.status_code not in (200, 201, 202):
    print(f"Submit failed: HTTP {resp.status_code}: {resp.text[:300]}")
    sys.exit(1)

job_id = resp.json().get("jobId") or resp.json().get("id")
print(f"  Job ID: {job_id}  — polling...")

terminal = {"COMPLETED", "FAILED", "COMPLETED_WITH_ERRORS", "PARTIAL_COMPLETED", "PARTIAL_SUCCESS"}
deadline = time.time() + 3600
dots = 0
import_status = "UNKNOWN"
while time.time() < deadline:
    r4 = requests.get(f"{ORG_URL}/data360/observable/v1/jobs/{job_id}",
        headers=H, timeout=30)
    data = r4.json()
    import_status = data.get("status", "UNKNOWN")
    if import_status in terminal:
        print(f"\r  {import_status}          ")
        if import_status == "FAILED":
            print(f"  Detail: {json.dumps(data.get('errors', data.get('detail', '')))[:500]}")
            print("Import FAILED — aborting before link phase.")
            sys.exit(1)
        if import_status in ("PARTIAL_COMPLETED", "PARTIAL_SUCCESS"):
            print("  Note: PARTIAL_COMPLETED is expected when some rules have no matching column.")
            print("  All successfully imported occurrences will be linked in Phase 3b.")
        break
    elapsed = int(time.time() - (deadline - 3600))
    print(f"\r  {import_status} ({elapsed}s){'.' * (dots % 4)}   ", end="", flush=True)
    dots += 1
    time.sleep(5)

# ══ PHASE 3b: PATCH template->occurrence relationships ════════════════════════

print("\nPhase 3b — Linking templates to occurrences...")
print(f"  {'Occurrence':<14} {'Template':<12} {'Status':<12} Detail")
print(f"  {'-' * 70}")

linked  = 0
skipped = 0
failed  = 0
errors  = []

for occ in occurrences:
    occ_id = occ["ref_id"]
    tpl_id = occ["tpl_ref"]
    url  = f"{ORG_URL}/data360/content/v1/assets/{tpl_id}?scheme=external"
    body = [{
        "operation": "add",
        "segment": "relationship",
        "items": [{
            "fromExternalIdentity": tpl_id,
            "toExternalIdentity":   occ_id,
            "association":          ASSOC_TYPE
        }]
    }]
    r5 = requests.patch(url, headers=H, json=body, timeout=30)

    if r5.status_code in (200, 201, 204):
        print(f"  {occ_id:<12} {tpl_id:<12} LINKED")
        linked += 1
    elif r5.status_code == 409:
        print(f"  {occ_id:<12} {tpl_id:<12} SKIP         already linked")
        skipped += 1
    else:
        print(f"  {occ_id:<12} {tpl_id:<12} FAILED       HTTP {r5.status_code}: {r5.text[:60]}")
        failed += 1
        errors.append((tpl_id, occ_id, r5.status_code, r5.text[:200]))

    time.sleep(0.3)

print(f"\n{'='*60}")
print(f"Import:   {import_status}")
print(f"Linked:   {linked} | Already linked: {skipped} | Failed: {failed}")
print(f"File:     {OCCURRENCE_FILE}")

if errors:
    print("\nFailed links:")
    for tpl, occ_id, code, txt in errors:
        print(f"  {tpl} -> {occ_id}  HTTP {code}: {txt}")
else:
    print("""
All phases complete.

Next step: Run MCC scan
  MCC -> Catalog Sources -> your catalog source -> Run
  Confirm Data Quality capability is enabled (Edit -> Capabilities -> Data Quality checked)

After scan: verify scores appeared
  python3 check_dq_links.py   -- spot-check one template
  python3 audit_dq_links.py   -- audit all templates vs File 15
""")
