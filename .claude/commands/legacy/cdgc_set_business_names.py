#!/usr/bin/env python3
"""
cdgc_set_business_names.py — Set Business Name on all governed column assets.

For every column that has a glossary term in Glossaries: Accepted, sets
Automatic Assignment = Enabled and reimports the Technical Data element sheet.
The platform then propagates the term name → Business Name automatically.

Requires CDGC_Columns_Export.xlsx in ~/Downloads (run cdgc_export_columns.py first
if it doesn't exist or is stale).

Usage:
  python3 ~/Documents/CDGC/cdgc_set_business_names.py
"""
import getpass
import time
from pathlib import Path

import requests

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip3 install openpyxl")

LOGIN_URL   = "https://dmp-us.informaticacloud.com"
ORG_URL     = "https://idmc-api.dmp-us.informaticacloud.com"
SOURCE_XLSX = Path.home() / "Downloads" / "CDGC_Columns_Export.xlsx"
OUT_XLSX    = Path.home() / "Downloads" / "CDGC_Set_BusinessNames.xlsx"
COLUMN_CLASS = "com.infa.odin.models.relational.Column"

# ── Auth ──────────────────────────────────────────────────────────────────────
username = input("IDMC Username: ")
password = getpass.getpass("IDMC Password: ")

resp = requests.post(f"{LOGIN_URL}/identity-service/api/v1/Login",
    json={"username": username, "password": password}, timeout=30)
resp.raise_for_status()
data = resp.json()
session_id, org_id = data["sessionId"], data["orgId"]

resp = requests.get(
    f"{LOGIN_URL}/identity-service/api/v1/jwt/Token?client_id=idmc_api&nonce=1234",
    headers={"IDS-SESSION-ID": session_id},
    cookies={"USER_SESSION": session_id}, timeout=30)
resp.raise_for_status()
jwt = (resp.json().get("token") or resp.json().get("jwt_token")
       or resp.json().get("access_token"))
print("✓ Authenticated\n")

H = {"Authorization": f"Bearer {jwt}", "X-INFA-ORG-ID": org_id}

# ── Load export workbook ───────────────────────────────────────────────────────
if not SOURCE_XLSX.exists():
    raise SystemExit(f"Export file not found: {SOURCE_XLSX}\nRun cdgc_export_columns.py first.")

print(f"Loading export: {SOURCE_XLSX}")
wb = openpyxl.load_workbook(SOURCE_XLSX)
ws = wb["Technical Data element"]

headers     = [c.value for c in ws[1]]
idx         = {h: i for i, h in enumerate(headers)}
auto_idx    = idx["Automatic Assignment"]
name_idx    = idx["Name"]
accepted_idx = idx["Glossaries: Accepted"]
bn_idx      = idx["Business Name"]

# ── Set Automatic Assignment = Enabled on all rows with a glossary term ────────
print("Scanning rows...")
modified = 0
skipped  = 0
rows_with_term = []

for row in ws.iter_rows(min_row=2):
    col_name = row[name_idx].value
    accepted = row[accepted_idx].value or ""
    bn       = row[bn_idx].value or ""

    if accepted:
        row[auto_idx].value = "Enabled"
        modified += 1
        rows_with_term.append((col_name, accepted, bn))
    else:
        skipped += 1

print(f"  {modified} rows set to Automatic Assignment = Enabled")
print(f"  {skipped} rows skipped (no glossary term)")
print()

# Show what will be updated
print(f"{'Column':<28} {'Glossaries: Accepted':<40} {'Business Name (before)'}")
print("-" * 100)
for col_name, accepted, bn in rows_with_term:
    print(f"  {col_name:<26} {accepted:<40} {bn or '(empty)'}")

print()
confirm = input(f"Update {modified} columns via reimport? (yes/no): ").strip().lower()
if confirm not in ("yes", "y"):
    print("Cancelled.")
    raise SystemExit(0)

# ── Save the modified workbook ─────────────────────────────────────────────────
# Keep only the "Technical Data element" sheet — the import only needs that sheet
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Technical Data element"

for row in ws.iter_rows(values_only=True):
    ws_out.append(list(row))

wb_out.save(OUT_XLSX)
print(f"\nSaved modified workbook: {OUT_XLSX}  ({modified} rows with Automatic Assignment = Enabled)")

# ── Import ─────────────────────────────────────────────────────────────────────
print("\nImporting via bulk import API...")
with open(OUT_XLSX, "rb") as f:
    files = {
        "file": (OUT_XLSX.name, f,
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "config": (None, '{"validationPolicy":"CONTINUE_ON_ERROR_WARNING"}',
                   "application/json"),
    }
    r_imp = requests.post(
        f"{ORG_URL}/data360/content/import/v1/assets",
        headers=H, files=files, timeout=60)

print(f"  HTTP {r_imp.status_code}")
if r_imp.status_code not in (200, 201, 202):
    print(f"  Error: {r_imp.text[:500]}")
    raise SystemExit(1)

job_id = r_imp.json().get("jobId") or r_imp.json().get("id")
print(f"  jobId: {job_id}")

# ── Poll ───────────────────────────────────────────────────────────────────────
print("\nPolling job status...")
poll_url = f"{ORG_URL}/data360/observable/v1/jobs/{job_id}"
terminal = {"COMPLETED", "FAILED", "COMPLETED_WITH_ERRORS", "PARTIAL_COMPLETED"}
final_status = None
for attempt in range(40):
    time.sleep(4)
    rp = requests.get(poll_url, headers=H, timeout=30)
    status = rp.json().get("status", rp.json().get("jobStatus", ""))
    print(f"  [{attempt+1}] {status}")
    if status in terminal:
        final_status = status
        details = rp.json().get("details") or rp.json().get("message") or ""
        if details:
            print(f"  details: {str(details)[:400]}")
        break
else:
    print("  Timed out waiting for import job")

# ── Verify via search API ──────────────────────────────────────────────────────
print(f"\nWaiting 15s for CDGC to process Business Names...")
time.sleep(15)

print("Verifying Business Names via search API...")
verify_cols = [name for name, _, _ in rows_with_term]
unique_cols = list(dict.fromkeys(verify_cols))  # dedupe, preserve order

populated = 0
still_empty = 0
results = []

for col_name in unique_cols:
    r_v = requests.post(
        f"{ORG_URL}/data360/search/v1/assets?knowledgeQuery={col_name}&segments=summary",
        headers={**H, "Content-Type": "application/json", "Accept": "application/json"},
        json={"from": 0, "size": 10,
              "filterSpec": [{"type": "simple", "attribute": "core.classType",
                              "values": [COLUMN_CLASS]}]}, timeout=30)
    hits = r_v.json().get("hits", [])
    for h in hits:
        s = h.get("summary", {})
        if s.get("core.name") == col_name:
            bn = s.get("core.businessName", "")
            results.append((col_name, bn))
            if bn:
                populated += 1
            else:
                still_empty += 1
            break
    time.sleep(0.3)

print(f"\n=== RESULTS ===\n")
print(f"{'Column':<28} {'Business Name'}")
print("-" * 60)
for col_name, bn in results:
    mark = "✓" if bn else "✗"
    print(f"  {mark} {col_name:<26} {bn or '(still empty — check UI)'}")

print(f"\nSummary: {populated} populated | {still_empty} still empty")
print(f"Import job status: {final_status}")

if still_empty:
    print("\nNote: The API search may lag. Check CDGC UI for the empty columns —")
    print("Business Name may already be visible there even if search hasn't indexed it yet.")
