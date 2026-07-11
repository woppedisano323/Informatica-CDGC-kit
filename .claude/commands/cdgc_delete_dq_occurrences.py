#!/usr/bin/env python3
"""
cdgc_delete_dq_occurrences.py

Deletes DQ Rule Occurrences from CDGC by Reference ID via the REST API.
Use this to clean up occurrences before re-importing with Operation=Create.

Reads reference IDs from the occurrence file if it exists; otherwise prompts
for a prefix and count to generate them.

Usage:
  python3 cdgc_delete_dq_occurrences.py
"""
import getpass
import time
import requests
from pathlib import Path

LOGIN_URL = "https://dmp-us.informaticacloud.com"
ORG_URL   = "https://idmc-api.dmp-us.informaticacloud.com"

# ── Import directory and reference ID resolution ──────────────────────────────
import_dir_raw = input("Import directory (e.g. ~/Downloads/CDGC_Import_MyClient): ").strip()
IMPORT_DIR = Path(import_dir_raw).expanduser()
OCC_FILE = IMPORT_DIR / "15_DQ_Rule_Occurrence.xlsx"

REF_IDS = []
if OCC_FILE.exists():
    from openpyxl import load_workbook
    wb = load_workbook(OCC_FILE)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ref_col = next((i for i, h in enumerate(hdr) if h and "Reference" in str(h)), 0)
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[ref_col]
        if val:
            REF_IDS.append(str(val).strip())
    print(f"Loaded {len(REF_IDS)} reference IDs from {OCC_FILE.name}")
else:
    occ_prefix = input("Occurrence prefix (e.g. DQOCC): ").strip()
    occ_count  = int(input("Number of occurrences to delete: ").strip())
    REF_IDS = [f"{occ_prefix}-{n}" for n in range(1, occ_count + 1)]
    print(f"Generated {len(REF_IDS)} reference IDs from prefix")

# ── Auth ──────────────────────────────────────────────────────────────────────
username = input("IDMC Username: ").strip()
password = getpass.getpass("IDMC Password: ")

r = requests.post(f"{LOGIN_URL}/identity-service/api/v1/Login",
    json={"username": username, "password": password}, timeout=30)
r.raise_for_status()
sid, oid = r.json()["sessionId"], r.json()["orgId"]
r2 = requests.get(
    f"{LOGIN_URL}/identity-service/api/v1/jwt/Token?client_id=idmc_api&nonce=1234",
    headers={"IDS-SESSION-ID": sid}, cookies={"USER_SESSION": sid}, timeout=30)
jwt = r2.json().get("token") or r2.json().get("jwt_token")
H = {"Authorization": f"Bearer {jwt}", "X-INFA-ORG-ID": oid}
print("✓ Authenticated\n")

# ── Delete each occurrence by external Reference ID ───────────────────────────
deleted = 0
not_found = 0
errors = 0

print(f"Deleting {len(REF_IDS)} occurrences...")
for ref_id in REF_IDS:
    r = requests.delete(
        f"{ORG_URL}/data360/content/v1/assets/{ref_id}?scheme=external",
        headers=H, timeout=30)
    if r.status_code in (200, 201):
        print(f"  ✓ Deleted: {ref_id}")
        deleted += 1
    elif r.status_code == 404:
        print(f"  - Not found (skipped): {ref_id}")
        not_found += 1
    else:
        print(f"  ✗ Error {r.status_code} on {ref_id}: {r.text[:100]}")
        errors += 1
    time.sleep(0.1)

print(f"""
{'═'*50}
  Deleted:    {deleted}
  Not found:  {not_found}  (never created — skipped)
  Errors:     {errors}
{'═'*50}

Next step — re-import clean:
  Upload {OCC_FILE.name} using the bulk import tool or API.
""")
