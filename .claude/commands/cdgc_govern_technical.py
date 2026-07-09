#!/usr/bin/env python3
"""
cdgc_govern_technical.py — Technical governance pipeline

Two steps:

  --step 1   Find gaps: fetch ungoverned columns, fuzzy-match to existing
             Business Terms, draft new terms for unmatched columns.
             Writes a review workbook for human approval.

  --step 2   Apply approvals: PATCH accepted links, import new terms,
             enable Automatic Assignment so Business Names propagate.

Flags:
  --company   Customer name (required, e.g. "First Capital Bank")
  --prefix    Asset prefix for new term IDs (required, e.g. FCB → FCBBT-N)
  --email     Governance owner email for new term imports (required for step 2)
  --workbook  Override workbook path (default: ~/Downloads/CDGC_Gap_Review_<slug>.xlsx)

Usage:
  python3 ~/Documents/CDGC/cdgc_govern_technical.py --step 1 --company "First Capital Bank" --prefix FCB
  python3 ~/Documents/CDGC/cdgc_govern_technical.py --step 2 --company "First Capital Bank" --prefix FCB --email you@example.com
"""
import argparse
import getpass
import re
import subprocess
import time
from pathlib import Path

import requests

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("pip3 install openpyxl")

# ── Config ────────────────────────────────────────────────────────────────────
LOGIN_URL    = "https://dmp-us.informaticacloud.com"
ORG_URL      = "https://idmc-api.dmp-us.informaticacloud.com"
COLUMN_CLASS = "com.infa.odin.models.relational.Column"
TERM_CLASS   = "com.infa.ccgf.models.governance.BusinessTerm"
SUBDOM_CLASS = "com.infa.ccgf.models.governance.Subdomain"
CONFIDENCE_THRESHOLD = 70

# ── Args ──────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--step",     choices=["1", "2"], help="Run step 1 (gap analysis) or step 2 (apply approvals)")
parser.add_argument("--company",  default="",         help="Customer name (required)")
parser.add_argument("--prefix",   default="",         help="Asset prefix for new term IDs, e.g. FCB → FCBBT-N")
parser.add_argument("--email",    default="",         help="Governance owner email for new terms (step 2)")
parser.add_argument("--workbook", default="",         help="Override workbook path")
args = parser.parse_args()

if not args.step:
    parser.print_help()
    raise SystemExit(0)

if not args.company or not args.prefix:
    print("ERROR: --company and --prefix are required.")
    print("  Example: python3 cdgc_govern_technical.py --step 1 --company \"First Capital Bank\" --prefix FCB")
    raise SystemExit(1)

slug    = args.company.strip().replace(" ", "_")
WB_PATH = Path(args.workbook) if args.workbook else Path.home() / "Downloads" / f"CDGC_Gap_Review_{slug}.xlsx"
IMP_PATH = WB_PATH.parent / f"CDGC_New_Terms_{slug}.xlsx"
BN_PATH  = WB_PATH.parent / f"CDGC_Set_BusinessNames_{slug}.xlsx"

# ── Auth ──────────────────────────────────────────────────────────────────────
username = input("IDMC Username: ")
password = getpass.getpass("IDMC Password: ")

resp = requests.post(f"{LOGIN_URL}/identity-service/api/v1/Login",
    json={"username": username, "password": password}, timeout=30)
resp.raise_for_status()
session_id = resp.json()["sessionId"]
org_id     = resp.json()["orgId"]

resp = requests.get(
    f"{LOGIN_URL}/identity-service/api/v1/jwt/Token?client_id=idmc_api&nonce=1234",
    headers={"IDS-SESSION-ID": session_id},
    cookies={"USER_SESSION": session_id}, timeout=30)
resp.raise_for_status()
jwt = (resp.json().get("token") or resp.json().get("jwt_token")
       or resp.json().get("access_token"))
print("✓ Authenticated\n")

H  = {"Authorization": f"Bearer {jwt}", "X-INFA-ORG-ID": org_id,
      "Content-Type": "application/json", "Accept": "application/json"}
HD = {"Authorization": f"Bearer {jwt}", "X-INFA-ORG-ID": org_id}

# ── API helpers ───────────────────────────────────────────────────────────────
def fetch_all(class_type, segments="summary"):
    results, offset = [], 0
    seen_ids = set()
    while True:
        r = requests.post(
            f"{ORG_URL}/data360/search/v1/assets?knowledgeQuery=*&segments={segments}",
            headers=H,
            json={"from": offset, "size": 100,
                  "filterSpec": [{"type": "simple", "attribute": "core.classType",
                                  "values": [class_type]}]}, timeout=30)
        r.raise_for_status()
        hits = r.json().get("hits", [])
        for h in hits:
            uid = h.get("core.identity", "")
            if uid and uid not in seen_ids:
                seen_ids.add(uid)
                results.append(h)
        if len(hits) < 100:
            break
        offset += 100
        time.sleep(0.2)
    return results


def patch_glossary(col_int_id, bt_ext_id):
    url  = f"{ORG_URL}/data360/content/v1/assets/{col_int_id}?scheme=internal"
    body = [{"operation": "add", "segment": "glossary",
             "items": [{"core.externalId": bt_ext_id}]}]
    r = requests.patch(url, headers=H, json=body, timeout=90)
    return r.status_code, r.text


def poll_job(job_id, label="job"):
    url      = f"{ORG_URL}/data360/observable/v1/jobs/{job_id}"
    terminal = {"COMPLETED", "FAILED", "COMPLETED_WITH_ERRORS", "PARTIAL_COMPLETED"}
    for attempt in range(40):
        time.sleep(4)
        rp     = requests.get(url, headers=H, timeout=30)
        status = rp.json().get("status", rp.json().get("jobStatus", ""))
        print(f"  [{attempt+1}] {status}")
        if status in terminal:
            details = rp.json().get("details") or rp.json().get("message") or ""
            if details:
                print(f"  details: {str(details)[:300]}")
            return status
    print(f"  Timed out waiting for {label}")
    return "TIMEOUT"


def import_xlsx(path, label="import"):
    print(f"Submitting {label}...")
    with open(path, "rb") as f:
        files = {
            "file": (path.name, f,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            "config": (None, '{"validationPolicy":"CONTINUE_ON_ERROR_WARNING"}',
                       "application/json"),
        }
        r = requests.post(f"{ORG_URL}/data360/content/import/v1/assets",
                          headers=HD, files=files, timeout=60)
    print(f"  HTTP {r.status_code}")
    if r.status_code not in (200, 201, 202):
        print(f"  Error: {r.text[:400]}")
        return None
    job_id = r.json().get("jobId") or r.json().get("id")
    print(f"  jobId: {job_id}")
    return job_id


# ── Excel styles ──────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1B2A4A")
HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
YES_FILL  = PatternFill("solid", fgColor="E8F5E9")
NEW_FILL  = PatternFill("solid", fgColor="E3F2FD")
DONE_FILL = PatternFill("solid", fgColor="F1F8E9")
WARN_FILL = PatternFill("solid", fgColor="FFF8E1")
YES_FONT  = Font(color="2E7D32", bold=True)
BLUE_FONT = Font(color="1565C0", bold=True)
THIN      = Side(style="thin", color="DDE1E8")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, cols):
    ws.row_dimensions[1].height = 22
    for i, (label, width) in enumerate(cols, 1):
        c = ws.cell(1, i, label)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

def write_row(ws, rn, values, fill=None, col1_font=None):
    for i, val in enumerate(values, 1):
        c = ws.cell(rn, i, val)
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=(i == len(values) - 1))
        if fill:
            c.fill = fill
    if col1_font:
        ws.cell(rn, 1).font = col1_font
    ws.row_dimensions[rn].height = 18

def inst_row(ws, text, ncols):
    ws.insert_rows(2)
    c = ws.cell(2, 1, text)
    c.font = Font(color="1565C0", italic=True, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


# ── Fuzzy matcher ─────────────────────────────────────────────────────────────
NOISE = {"of","the","a","an","and","or","for","in","on","at","to","is","as",
         "by","be","its","per","from","with","after","before"}

def tokens(s):
    return set(w.lower() for w in re.split(r'[_\s]+', s) if w)

def fuzzy_score(col_name, term_name, term_desc=""):
    ct = tokens(col_name)
    tt = tokens(term_name)
    if not ct or not tt:
        return 0
    if ct == tt:
        return 95
    if ct.issubset(tt):
        return 85
    if tt.issubset(ct):
        return 80
    overlap   = ct & tt
    meaningful = overlap - NOISE
    if not meaningful:
        return 0
    jaccard    = len(overlap) / len(ct | tt)
    desc_boost = 0
    if term_desc:
        dt = tokens(term_desc)
        if len(ct & dt - NOISE) >= len(ct - NOISE):
            desc_boost = 15
        elif ct & dt - NOISE:
            desc_boost = 8
    return min(100, int(jaccard * 75) + desc_boost)

def best_match(col_name, terms):
    best = None
    for t in terms:
        s = fuzzy_score(col_name, t["name"], t.get("description", ""))
        if s >= CONFIDENCE_THRESHOLD:
            if best is None or s > best[2]:
                best = (t["ext_id"], t["name"], s)
    return best


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Find gaps, fuzzy match, generate review workbook
# ══════════════════════════════════════════════════════════════════════════════
def run_step_1():
    print("=" * 60)
    print("STEP 1 — Gap analysis → review workbook")
    print("=" * 60)

    # Fetch columns with selfAttributes — core.businessName is populated there after
    # MCC AUTO ACCEPTS a Glossary Association. The search API glossary segment returns
    # null for Column assets (known API limitation) so selfAttributes is the only
    # reliable way to detect MCC-linked columns.
    print("\nFetching scanned columns (with selfAttributes for businessName)...")
    col_hits = fetch_all(COLUMN_CLASS, segments="summary,selfAttributes")
    print(f"  {len(col_hits)} columns found")

    print("Fetching Business Terms...")
    term_hits = fetch_all(TERM_CLASS)
    terms = []
    for h in term_hits:
        s      = h.get("summary") or {}
        ext_id = h.get("core.externalId") or s.get("core.externalId", "")
        name   = s.get("core.name", "")
        desc   = s.get("core.description", "")
        if ext_id and name and not ext_id.startswith("BT-"):
            terms.append({"ext_id": ext_id, "name": name, "description": desc})
    print(f"  {len(terms)} customer-authored terms")

    print("Fetching Subdomains...")
    subdom_hits = fetch_all(SUBDOM_CLASS)
    subdom_lookup = {}
    for h in subdom_hits:
        s      = h.get("summary") or {}
        ext_id = h.get("core.externalId") or s.get("core.externalId", "")
        name   = s.get("core.name", "")
        if ext_id and name:
            subdom_lookup[name] = ext_id
    print(f"  {len(subdom_lookup)} subdomains\n")

    # Categorize columns
    # MCC puts its Glossary Association matches in "recommended", not "accepted".
    # "accepted" only appears after Automatic Assignment runs or manual acceptance.
    # We treat both as already governed — no action needed for either.
    already_governed = []   # MCC recommended or accepted a term
    suggest_link     = []   # no term at all, but fuzzy match found
    suggest_new      = []   # no term at all, no match — needs new Business Term

    for h in col_hits:
        s            = h.get("summary") or {}
        sa           = h.get("selfAttributes") or {}
        name         = s.get("core.name", "")
        loc          = s.get("core.location", "")
        int_id       = h.get("core.identity", "")
        parts        = loc.rstrip("/").split("/")
        table        = parts[-2] if len(parts) >= 2 else ""
        business_name = sa.get("core.businessName", "") or ""

        if not name:
            continue

        if business_name:
            # MCC AUTO ACCEPTED a term — businessName is populated
            already_governed.append({
                "name": name, "table": table, "id": int_id,
                "term_name": business_name,
                "term_ext":  "",
                "status": "MCC Accepted",
            })
        else:
            match = best_match(name, terms)
            if match:
                suggest_link.append({
                    "name": name, "table": table, "id": int_id,
                    "term_ext": match[0], "term_name": match[1], "confidence": match[2],
                })
            else:
                draft_name = name.replace("_", " ").title()
                suggest_new.append({
                    "name": name, "table": table, "id": int_id,
                    "draft_name": draft_name,
                    "draft_desc": f"[Edit: description for {draft_name}]",
                    "subdomain": "",
                    "domain": "",
                })

    print(f"Results:")
    print(f"  Already governed (MCC) : {len(already_governed)}")
    print(f"  Suggested links        : {len(suggest_link)}")
    print(f"  Suggested new terms    : {len(suggest_new)}")
    print()

    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = Workbook()

    # Tab 1: Suggested Links
    ws1 = wb.active
    ws1.title = "Suggested Links"
    ws1.freeze_panes = "A3"
    COLS1 = [("APPROVE",14),("Table",22),("Column Name",22),
             ("Suggested Business Term",28),("Confidence",12),
             ("Term Description",42),("Term ext_id",14)]
    style_header(ws1, COLS1)
    inst_row(ws1, "→ Set APPROVE to YES for each match you accept, then run Step 2.", len(COLS1))
    for row in sorted(suggest_link, key=lambda r: (-r["confidence"], r["table"], r["name"])):
        desc = next((t["description"] for t in terms if t["ext_id"] == row["term_ext"]), "")
        fill = YES_FILL if row["confidence"] >= 85 else WARN_FILL
        write_row(ws1, ws1.max_row+1,
                  ["", row["table"], row["name"], row["term_name"],
                   f"{row['confidence']}%", desc, row["term_ext"]],
                  fill=fill, col1_font=YES_FONT)
    if not suggest_link:
        ws1.cell(ws1.max_row+1, 1, "No suggested links found.")

    # Tab 2: Suggested New Terms
    ws2 = wb.create_sheet("Suggested New Terms")
    ws2.freeze_panes = "A3"
    COLS2 = [("APPROVE",14),("Table",22),("Column Name",22),
             ("Business Term Name",28),("Description",52),
             ("Subdomain",28),("Domain",22)]
    style_header(ws2, COLS2)
    inst_row(ws2, "→ Edit Name/Description/Subdomain, set APPROVE=YES. Run Step 2 when done.", len(COLS2))
    for row in sorted(suggest_new, key=lambda r: (r["table"], r["name"])):
        write_row(ws2, ws2.max_row+1,
                  ["", row["table"], row["name"], row["draft_name"],
                   row["draft_desc"], row["subdomain"], row["domain"]],
                  fill=NEW_FILL, col1_font=BLUE_FONT)
    if not suggest_new:
        ws2.cell(ws2.max_row+1, 1, "No new terms needed — all columns matched existing terms.")

    # Tab 3: Already Governed
    ws3 = wb.create_sheet("Already Governed")
    ws3.freeze_panes = "A2"
    COLS3 = [("Table",22),("Column Name",22),("Business Term",28),
             ("Term ext_id",14),("Status",18)]
    style_header(ws3, COLS3)
    for row in sorted(already_governed, key=lambda r: (r["table"], r["name"])):
        write_row(ws3, ws3.max_row+1,
                  [row["table"], row["name"], row["term_name"],
                   row["term_ext"], row["status"]],
                  fill=DONE_FILL)

    wb.save(WB_PATH)
    subprocess.Popen(["open", str(WB_PATH)])
    print(f"✓ Workbook written and opened: {WB_PATH}")
    print(f"""
{'═'*60}
HUMAN REVIEW REQUIRED
{'═'*60}

FILE: {WB_PATH.name}

─── TAB 1: Suggested Links ───────────────────────────────
  Unlinked columns fuzzy-matched to existing Business Terms.
  Green = high confidence (≥85%). Yellow = review carefully.
  Set APPROVE = YES to accept, leave blank to skip.

─── TAB 2: Suggested New Terms ───────────────────────────
  Columns with no matching term — AI-drafted names/descriptions.
  Edit any row before approving. Set APPROVE = YES to create
  the term in CDGC and link the column.

─── TAB 3: Already Governed ──────────────────────────────
  Columns MCC already linked to Business Terms. No action needed.

─── WHEN DONE ────────────────────────────────────────────
  1. Save the workbook (Cmd+S)
  2. Run Step 2:

     python3 ~/Documents/CDGC/cdgc_govern_technical.py \\
       --step 2 --company "{args.company}" --prefix {args.prefix} \\
       --email your@email.com

{'═'*60}
""")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Apply approvals: link columns, import new terms, set Business Names
# ══════════════════════════════════════════════════════════════════════════════
def run_step_2():
    print("=" * 60)
    print("STEP 2 — Apply approvals")
    print("=" * 60)

    if not WB_PATH.exists():
        raise SystemExit(f"ERROR: Workbook not found: {WB_PATH}\nRun --step 1 first.")

    wb = load_workbook(WB_PATH)

    # Fetch all columns once — keyed by internal identity
    print("\nFetching scanned columns...")
    col_hits = fetch_all(COLUMN_CLASS, segments="summary,selfAttributes")
    col_map = {}
    for h in col_hits:
        s      = h.get("summary") or {}
        name   = s.get("core.name", "")
        int_id = h.get("core.identity", "")
        ext_id = h.get("core.externalId") or s.get("core.externalId", "")
        if name and int_id:
            col_map[name.upper()] = {"int_id": int_id, "ext_id": ext_id}
    print(f"  {len(col_map)} columns found")

    # Build term name → ext_id lookup for Part D
    all_terms_for_lookup = fetch_all(TERM_CLASS)
    term_name_to_ext = {}
    for h in all_terms_for_lookup:
        s      = h.get("summary") or {}
        ext_id = h.get("core.externalId") or s.get("core.externalId", "")
        name   = s.get("core.name", "")
        if name and ext_id and not ext_id.startswith("BT-"):
            term_name_to_ext[name] = ext_id

    # ── Part A: Link approved Tab 1 rows to existing terms ────────────────────
    ok_a = fail_a = skip_a = 0
    linked_col_ext_ids = []  # track for Automatic Assignment step

    if "Suggested Links" in wb.sheetnames:
        ws1  = wb["Suggested Links"]
        hdrs = [c.value for c in ws1[1]]
        idx  = {h: i for i, h in enumerate(hdrs)}
        print(f"\nPart A — Linking approved columns to existing terms...")
        for row in ws1.iter_rows(min_row=3, values_only=True):
            if str(row[idx.get("APPROVE", 0)] or "").strip().upper() != "YES":
                skip_a += 1
                continue
            term_ext = row[idx["Term ext_id"]]
            col_name = row[idx["Column Name"]]
            table    = row[idx["Table"]]
            if not term_ext:
                skip_a += 1
                continue
            col = col_map.get(col_name.upper())
            if not col:
                print(f"  ✗ {table}.{col_name} — not found in scan")
                fail_a += 1
                continue
            status, body = patch_glossary(col["int_id"], term_ext)
            already = status == 500 and term_ext in body
            if status in (200, 204) or already:
                tag = " (already linked)" if already else ""
                print(f"  ✓ {table}.{col_name} → {term_ext}{tag}")
                ok_a += 1
                linked_col_ext_ids.append(col["ext_id"])
            else:
                print(f"  ✗ {table}.{col_name}  [{status}] {body[:80]}")
                fail_a += 1
            time.sleep(0.3)
        print(f"  Done: {ok_a} linked, {fail_a} failed, {skip_a} skipped")

    # ── Part B: Import approved new terms ─────────────────────────────────────
    approved_new = []
    if "Suggested New Terms" in wb.sheetnames:
        ws2  = wb["Suggested New Terms"]
        hdrs = [c.value for c in ws2[1]]
        idx2 = {h: i for i, h in enumerate(hdrs)}
        for row in ws2.iter_rows(min_row=3, values_only=True):
            if str(row[idx2.get("APPROVE", 0)] or "").strip().upper() != "YES":
                continue
            approved_new.append({
                "col_name":  row[idx2["Column Name"]],
                "table":     row[idx2["Table"]],
                "term_name": row[idx2["Business Term Name"]],
                "term_desc": row[idx2["Description"]],
                "subdomain": row[idx2["Subdomain"]] or "",
                "domain":    row[idx2["Domain"]] or "",
            })

    term_ref_map = {}  # term_name → ext_id for linking step

    if approved_new:
        gov_email = args.email.strip()
        if not gov_email:
            gov_email = input("\nGovernance owner email for new terms: ").strip()

        print(f"\nPart B — Importing {len(set(r['term_name'] for r in approved_new))} new Business Terms...")

        # Find next available Reference ID
        all_terms = fetch_all(TERM_CLASS)
        prefix_bt = f"{args.prefix}BT-"
        max_n = 0
        for h in all_terms:
            s      = h.get("summary") or {}
            ext_id = h.get("core.externalId") or s.get("core.externalId", "")
            if ext_id.startswith(prefix_bt):
                try:
                    n = int(ext_id[len(prefix_bt):])
                    max_n = max(max_n, n)
                except ValueError:
                    pass
        next_n = max_n + 1
        print(f"  Next Reference ID: {prefix_bt}{next_n}")

        # Fetch subdomains for parent reference
        subdom_hits = fetch_all(SUBDOM_CLASS)
        subdom_lookup = {}
        for h in subdom_hits:
            s      = h.get("summary") or {}
            ext_id = h.get("core.externalId") or s.get("core.externalId", "")
            name   = s.get("core.name", "")
            if ext_id and name:
                subdom_lookup[name] = ext_id

        # Existing terms — skip duplicates, link directly
        existing_term_map = {}
        for h in all_terms:
            s      = h.get("summary") or {}
            ext_id = h.get("core.externalId") or s.get("core.externalId", "")
            name   = s.get("core.name", "")
            if name and ext_id:
                existing_term_map[name] = ext_id

        seen_terms = {}
        skipped_no_subdomain = []
        for row in approved_new:
            tname = row["term_name"]
            if tname in existing_term_map:
                term_ref_map[tname] = existing_term_map[tname]
            elif not row["subdomain"].strip():
                # Guard: blank Subdomain would import the term at Top Level with no parent.
                # Require the reviewer to fill in Subdomain before approving.
                skipped_no_subdomain.append(tname)
            elif tname not in seen_terms:
                seen_terms[tname] = row

        if skipped_no_subdomain:
            print(f"\n  ⚠  {len(skipped_no_subdomain)} approved term(s) skipped — Subdomain is blank:")
            for t in skipped_no_subdomain:
                print(f"       • {t}")
            print("     Fill in the Subdomain column in the workbook and re-run Step 2.\n")

        if existing_term_map and any(r["term_name"] in existing_term_map for r in approved_new):
            already_count = sum(1 for r in approved_new if r["term_name"] in existing_term_map)
            print(f"  {already_count} term(s) already exist — linking directly")

        # Build import workbook
        imp_wb = Workbook()
        ws_imp = imp_wb.active
        ws_imp.title = "Business Term"
        IMP_COLS = [
            ("Reference ID",16),("Name",30),("Description",55),
            ("Alias Names",20),("Business Logic",20),("Critical Data Element",22),
            ("Examples",20),("Format Type",14),("Format Description",20),
            ("Lifecycle",14),("Security Level",16),("Classifications",20),
            ("Reference Data",20),("Operation",12),
            ("Parent: Subdomain",35),("Parent: Business Term",20),
            ("Parent: Metric",20),("Parent: Domain",20),
            ("Stakeholder: Governance Owner",35),
            ("Stakeholder: Governance Administrator",35),
        ]
        style_header(ws_imp, IMP_COLS)

        for i, (name, row) in enumerate(seen_terms.items(), 0):
            ref_id     = f"{prefix_bt}{next_n + i}"
            subdom_ext = subdom_lookup.get(row["subdomain"], "")
            if not subdom_ext:
                # Subdomain name didn't match any known subdomain in the org —
                # warn and skip rather than importing with a broken parent reference.
                print(f"  ⚠  Skipping '{name}' — subdomain '{row['subdomain']}' not found in org")
                continue
            parent_ref = f"{row['subdomain']} | {subdom_ext}"
            term_ref_map[name] = ref_id
            write_row(ws_imp, ws_imp.max_row+1, [
                ref_id, name, row["term_desc"],
                "", "", "false", "", "Text", "", "Published", "Internal",
                "", "", "Create",
                parent_ref, "", "", "",
                gov_email, gov_email,
            ])

        imp_wb.save(IMP_PATH)
        print(f"  Saved: {IMP_PATH} ({len(seen_terms)} new terms)")

        job_id = import_xlsx(IMP_PATH, "new Business Terms")
        if job_id:
            status = poll_job(job_id, "Business Term import")
            print(f"  Import status: {status}")
        else:
            raise SystemExit("Import failed.")

        print("  Waiting 15s for indexing...")
        time.sleep(15)

        # ── Part C: Link approved Tab 2 columns to new terms ──────────────────
        print(f"\nPart C — Linking {len(approved_new)} column(s) to new terms...")
        ok_c = fail_c = 0
        for row in approved_new:
            col_name  = row["col_name"]
            table     = row["table"]
            term_name = row["term_name"]
            ext_id    = term_ref_map.get(term_name)
            if not ext_id:
                print(f"  ✗ {table}.{col_name} — no ref_id for {term_name!r}")
                fail_c += 1
                continue
            col = col_map.get(col_name.upper())
            if not col:
                print(f"  ✗ {table}.{col_name} — not found in scan")
                fail_c += 1
                continue
            status, body = patch_glossary(col["int_id"], ext_id)
            already = status == 500 and ext_id in body
            if status in (200, 204) or already:
                tag = " (already linked)" if already else ""
                print(f"  ✓ {table}.{col_name} → {term_name} ({ext_id}){tag}")
                ok_c += 1
                linked_col_ext_ids.append(col["ext_id"])
            else:
                print(f"  ✗ {table}.{col_name} → {ext_id}  [{status}] {body[:80]}")
                fail_c += 1
            time.sleep(0.3)
        print(f"  Done: {ok_c} linked, {fail_c} failed")

    # ── Part D: Business Names ─────────────────────────────────────────────────
    # Automatic Assignment via bulk import API does not work — hard FAILED on every attempt.
    # Business Names propagate correctly via MCC Glossary Association re-run (Option A).
    print("\nPart D — Business Names")
    print("  Column→Term links are now in place (Parts A/B/C above).")
    print("  To propagate Business Names to all linked columns:")
    print("  1. Go to MCC → FCB_Financial_Snowflake catalog source")
    print("  2. Open the Glossary Association task settings")
    print("  3. Set 'Assign Business Names and Descriptions' → Yes")
    print("  4. Disable 'Keep Existing Business Names and Descriptions'")
    print("  5. Re-execute the Glossary Association job")
    print("\n✓ Step 2 complete.\n")
    return

    print("  Exporting Technical Data elements from CDGC...")

    # Step 1: trigger export
    # Per Chapter 7 docs: body is {"size": 0} to export all matching assets.
    # filterSpec is a search API body param — not valid here; use knowledgeQuery instead.
    # Response returns jobId + outputURI — outputURI is the download path.
    export_r = requests.post(
        f"{ORG_URL}/data360/search/export/v1/assets"
        f"?knowledgeQuery=data+elements&segments=summary,glossary"
        f"&fileType=EXCEL&summaryViews=Technical",
        headers=H,
        json={"size": 0},
        timeout=60)
    export_r.raise_for_status()
    export_resp   = export_r.json()
    export_job_id = export_resp.get("jobId") or export_resp.get("id")
    output_uri    = export_resp.get("outputURI", "")
    print(f"  Export jobId: {export_job_id}")
    print(f"  Output URI:   {output_uri}")

    # Step 2: poll until COMPLETED then download via outputURI
    export_url = None
    for attempt in range(90):
        time.sleep(5)
        ep = requests.get(f"{ORG_URL}/data360/observable/v1/jobs/{export_job_id}",
                          headers=H, timeout=30)
        status = ep.json().get("status", "")
        print(f"  [{attempt+1}] {status}")
        if status == "COMPLETED":
            export_url = f"{ORG_URL}{output_uri}" if output_uri else \
                         f"{ORG_URL}/data360/observable/v1/jobs/{export_job_id}/outputProperties/files/Export_File"
            print(f"  Download URL: {export_url}")
            break
        if status in ("FAILED", "COMPLETED_WITH_ERRORS"):
            print(f"  Export failed: {ep.json()}")
            raise SystemExit("Export failed — cannot proceed with Automatic Assignment.")

    if not export_url:
        raise SystemExit("Export job timed out after 90 attempts.")

    # Step 3: download the exported xlsx
    dl = requests.get(export_url, headers=HD, timeout=120)
    dl.raise_for_status()
    EXPORT_PATH = BN_PATH.parent / f"CDGC_Export_Columns_{slug}.xlsx"
    EXPORT_PATH.write_bytes(dl.content)
    print(f"  Downloaded export: {EXPORT_PATH} ({len(dl.content):,} bytes)")

    # Step 4: open export, find Technical Data element sheet, set Automatic Assignment
    exp_wb = load_workbook(EXPORT_PATH)
    if "Technical Data element" not in exp_wb.sheetnames:
        print(f"  Available sheets: {exp_wb.sheetnames}")
        raise SystemExit("'Technical Data element' sheet not found in export.")

    ws_bn = exp_wb["Technical Data element"]
    hdrs_bn = [c.value for c in ws_bn[1]]
    print(f"  Export sheet headers: {hdrs_bn}")

    try:
        aa_idx   = hdrs_bn.index("Automatic Assignment") + 1
        gl_idx   = hdrs_bn.index("Glossaries: Accepted") + 1
        op_idx   = hdrs_bn.index("Operation") + 1
    except ValueError as e:
        raise SystemExit(f"Expected column not found in export: {e}")

    enabled = 0
    for row in ws_bn.iter_rows(min_row=2):
        glossary_val = row[gl_idx - 1].value or ""
        if glossary_val:
            row[aa_idx - 1].value = "Enabled"
            row[op_idx - 1].value = "Update"
            enabled += 1

    # Preserve all sheets — docs require the full exported file structure (Table of Contents,
    # Technical Data element, View Technical). Stripping sheets causes hard FAILED.
    exp_wb.save(BN_PATH)
    print(f"  Saved: {BN_PATH} ({enabled} rows with Automatic Assignment = Enabled)")

    if enabled == 0:
        print("  No governed columns found in export — skipping import.")
    else:
        job_id = import_xlsx(BN_PATH, "Automatic Assignment")
        if job_id:
            status = poll_job(job_id, "Automatic Assignment import")
            print(f"  Import status: {status}")
        else:
            raise SystemExit("Automatic Assignment import failed.")

        # Verify
        print("\nVerifying Business Names (waiting 20s for index)...")
        time.sleep(20)
        col_names = []
        for row in ws_bn.iter_rows(min_row=2, values_only=True):
            if row[gl_idx - 1]:
                name_val = row[hdrs_bn.index("Name")]
                if name_val and name_val not in col_names:
                    col_names.append(name_val)

        populated = 0
        missing   = []
        for col_name in col_names:
            rv = requests.post(
                f"{ORG_URL}/data360/search/v1/assets?knowledgeQuery=*&segments=summary",
                headers=H,
                json={"from": 0, "size": 10,
                      "filterSpec": [{"type": "simple", "attribute": "core.classType",
                                      "values": [COLUMN_CLASS]}],
                      "query": col_name}, timeout=30)
            bn = ""
            for h in rv.json().get("hits", []):
                s = h.get("summary", {})
                if s.get("core.name") == col_name:
                    bn = s.get("core.businessName", "") or ""
                    break
            if bn:
                populated += 1
            else:
                missing.append(col_name)
            time.sleep(0.2)

        print(f"\n  Business Names populated: {populated}/{len(col_names)}")
        if missing:
            print("  Not yet visible via API (check UI — index lag):")
            for n in missing:
                print(f"    • {n}")
        else:
            print(f"  ✓ All {populated} governed columns have Business Names.")

    print(f"\n✓ Step 2 complete.\n")


# ── Main ──────────────────────────────────────────────────────────────────────
if args.step == "1":
    run_step_1()
elif args.step == "2":
    run_step_2()
