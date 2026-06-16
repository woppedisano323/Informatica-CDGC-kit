#!/usr/bin/env python3
"""
cdgc_export_columns.py  —  Export all column assets from CDGC to Excel.

Downloads the exported file to ~/Downloads/CDGC_Columns_Export.xlsx
so we can examine the exact column names and Automatic Assignment field.

Usage:
  python3 ~/Documents/CDGC/cdgc_export_columns.py
"""
import getpass
import time
from pathlib import Path

import requests

LOGIN_URL = "https://dmp-us.informaticacloud.com"
ORG_URL   = "https://idmc-api.dmp-us.informaticacloud.com"
COLUMN_CLASS = "com.infa.odin.models.relational.Column"
OUT_PATH  = Path.home() / "Downloads" / "CDGC_Columns_Export.xlsx"

# ── Auth ──────────────────────────────────────────────────────────────────────
username = input("IDMC Username: ")
password = getpass.getpass("IDMC Password: ")

resp = requests.post(f"{LOGIN_URL}/identity-service/api/v1/Login",
    json={"username": username, "password": password}, timeout=30)
resp.raise_for_status()
data = resp.json()
session_id, org_id = data["sessionId"], data["orgId"]

resp = requests.get(
    f"{LOGIN_URL}/identity-service/api/v1/jwt/Token?client_id=idmc_api&nonce=1234",
    headers={"IDS-SESSION-ID": session_id},
    cookies={"USER_SESSION": session_id}, timeout=30)
resp.raise_for_status()
jwt = (resp.json().get("token") or resp.json().get("jwt_token")
       or resp.json().get("access_token"))
print("✓ Authenticated\n")

H = {"Authorization": f"Bearer {jwt}", "X-INFA-ORG-ID": org_id,
     "Content-Type": "application/json", "Accept": "application/json"}

# ── Trigger export job ────────────────────────────────────────────────────────
print("Triggering export job for all Column assets...")
export_url = (
    f"{ORG_URL}/data360/search/export/v1/assets"
    f"?knowledgeQuery=*"
    f"&segments=summary,glossary,selfAttributes"
    f"&fileName=CDGC_Columns_Export"
    f"&fileType=EXCEL"
    f"&summaryViews=Technical"
)
r = requests.post(export_url, headers=H,
                  json={"from": 0, "size": 200,
                        "filterSpec": [{"type": "simple",
                                        "attribute": "core.classType",
                                        "values": [COLUMN_CLASS]}]},
                  timeout=60)
print(f"  HTTP {r.status_code}")
if r.status_code not in (200, 201, 202):
    print(f"  Error: {r.text[:500]}")
    raise SystemExit(1)

job_data = r.json()
print(f"  jobId:       {job_data.get('jobId')}")
tracking = job_data.get("trackingURI", "")
output   = job_data.get("outputURI", "")
print(f"  trackingURI: {tracking}")
print(f"  outputURI:   {output}\n")

# ── Poll until complete ───────────────────────────────────────────────────────
if tracking:
    print("Polling job status...")
    poll_url = f"{ORG_URL}{tracking}" if not tracking.startswith("http") else tracking
    for attempt in range(30):
        time.sleep(3)
        rp = requests.get(poll_url, headers=H, timeout=30)
        if rp.status_code != 200:
            print(f"  Poll HTTP {rp.status_code} — {rp.text[:200]}")
            break
        status = rp.json().get("status", rp.json().get("jobStatus", ""))
        print(f"  [{attempt+1}] status: {status}")
        if status in ("COMPLETED", "SUCCESS", "FAILED", "ERROR"):
            # Try to get outputURI from the poll response if not in original
            if not output:
                output = (rp.json().get("outputURI", "") or
                          rp.json().get("outputProperties", {}).get("files", {}).get("Export_File", ""))
            break
    else:
        print("  Timed out waiting for export job")

# ── Download the file ─────────────────────────────────────────────────────────
if output:
    print(f"\nDownloading export file...")
    dl_url = f"{ORG_URL}{output}" if not output.startswith("http") else output
    rd = requests.get(dl_url, headers=H, timeout=60)
    print(f"  HTTP {rd.status_code}  content-type: {rd.headers.get('content-type','?')}")
    if rd.status_code == 200:
        OUT_PATH.write_bytes(rd.content)
        print(f"  Saved to: {OUT_PATH}  ({len(rd.content):,} bytes)")
    else:
        print(f"  Download failed: {rd.text[:300]}")
else:
    print("\nNo outputURI found — check the job in CDGC UI")
    raise SystemExit(1)

# ── Peek at the file ──────────────────────────────────────────────────────────
try:
    import openpyxl
    wb = openpyxl.load_workbook(OUT_PATH)
    print(f"\nSheets in workbook: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        headers = [c.value for c in ws[1]]
        print(f"\n── Sheet: {sheet_name!r}  ({ws.max_row-1} data rows) ──")
        print(f"   Columns: {[h for h in headers if h]}")
        print(f"\n   First 3 rows:")
        for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
            print(f"   {list(row)}")
except ImportError:
    print("\nopenpyxl not installed — run: pip3 install openpyxl")
    print(f"File saved at {OUT_PATH} — open it in Excel to inspect.")

print("\nDone.")
