#!/usr/bin/env python3
"""
cdgc_gap_analyzer.py  —  Governance Gap Analyzer

Three-phase workflow to close the governance gap:

  PHASE 1 — Analyze (default)
    Fetches scanned columns + existing Business Terms, produces a 3-tab review workbook:
      Tab 1 — Suggested Links    : ungoverned columns matched to EXISTING terms
      Tab 2 — Suggested New Terms: ungoverned columns with no match — AI-drafted definitions
      Tab 3 — Already Governed   : reference only

  PHASE 2 — Link existing terms (--apply)
    Reads Tab 1 APPROVE=YES rows → PATCHes column glossary link to existing term.
    Also generates 11_Business_Term_Gap.xlsx for CDGC bulk import (Tab 2 approved rows).
    Import that file via CDGC UI before running Phase 3.

  PHASE 3 — Link new terms (--link-new)
    Reads Tab 2 APPROVE=YES rows → searches CDGC for each term by name →
    PATCHes column glossary link. Run AFTER importing 11_Business_Term_Gap.xlsx.

Usage:
  python3 cdgc_gap_analyzer.py                                    # analyze
  python3 cdgc_gap_analyzer.py --company "Acme Corp" --prefix ACM # analyze, named output
  python3 cdgc_gap_analyzer.py --apply                            # phase 2
  python3 cdgc_gap_analyzer.py --link-new                         # phase 3
"""
import argparse
import getpass
import re
import subprocess
import time
from pathlib import Path

import requests

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  Run: pip3 install openpyxl")
    raise SystemExit(1)

# ── Config ────────────────────────────────────────────────────────────────────
LOGIN_URL = "https://dmp-us.informaticacloud.com"
ORG_URL   = "https://idmc-api.dmp-us.informaticacloud.com"

COLUMN_CLASS = "com.infa.odin.models.relational.Column"
TERM_CLASS   = "com.infa.ccgf.models.governance.BusinessTerm"
SUBDOM_CLASS = "com.infa.ccgf.models.governance.Subdomain"

CONFIDENCE_THRESHOLD = 70

# Already-linked columns — skip in gap analysis.
# IDs below are FSI Financial Services reference examples only — superseded by cdgc_govern_technical.py
# which resolves term IDs live from the org at runtime. Do not rely on these IDs.
ALREADY_LINKED_MAP = {
    "CUSTOMER_ID":            ("",  "Customer ID"),
    "SSN":                    ("",  "Social Security Number"),
    "DATE_OF_BIRTH":          ("",  "Date of Birth"),
    "EMAIL":                  ("",  "Email Address"),
    "PHONE_NUMBER":           ("",  "Phone Number"),
    "CREDIT_SCORE":           ("",  "Credit Score"),
    "TRANSACTION_ID":         ("",  "Transaction ID"),
    "AMOUNT":                 ("",  "Transaction Amount"),
    "TRANSACTION_DATE":       ("",  "Transaction Date"),
    "POSTING_DATE":           ("",  "Post Date"),
    "CURRENCY":               ("",  "Currency Code"),
    "ACCOUNT_CODE":           ("",  "GL Account Number"),
    "DEBIT_AMOUNT":           ("",  "Debit Amount"),
    "CREDIT_AMOUNT":          ("",  "Credit Amount"),
    "STATUS":                 ("",  "Entry Status"),
    "FISCAL_PERIOD":          ("",  "Accounting Period"),
    "PROBABILITY_OF_DEFAULT": ("",  "Probability of Default"),
    "LOSS_GIVEN_DEFAULT":     ("",  "Loss Given Default"),
}

# AI-drafted new Business Terms for ungoverned columns
NEW_TERM_DRAFTS = {
    # CUSTOMER_MASTER
    "FIRST_NAME":        ("First Name",               "Customer legal first name as provided at onboarding",                          "Customer Identity",    "Customer & KYC"),
    "LAST_NAME":         ("Last Name",                "Customer legal last name or family name",                                      "Customer Identity",    "Customer & KYC"),
    "ADDRESS_LINE1":     ("Street Address",           "Primary street address line for the customer's residence or mailing address",  "Customer Identity",    "Customer & KYC"),
    "CITY":              ("City",                     "City of residence or mailing address for the customer",                        "Customer Identity",    "Customer & KYC"),
    "STATE":             ("State",                    "US state abbreviation for the customer's address",                             "Customer Identity",    "Customer & KYC"),
    "ZIP_CODE":          ("Zip Code",                 "US postal zip code for the customer's address",                                "Customer Identity",    "Customer & KYC"),
    "ACCOUNT_STATUS":    ("Account Status",           "Current status of the customer account: ACTIVE, INACTIVE, SUSPENDED, CLOSED", "Customer Identity",    "Customer & KYC"),
    "CUSTOMER_SINCE":    ("Customer Since Date",      "Date the customer relationship was first established with the institution",    "Customer Identity",    "Customer & KYC"),
    "ANNUAL_INCOME":     ("Annual Income",            "Customer's declared annual income in USD, used for credit and risk assessment","Customer Identity",    "Customer & KYC"),
    "RISK_TIER":         ("Customer Risk Tier",       "Internal risk classification: LOW, MEDIUM, HIGH, VERY_HIGH",                  "KYC & Compliance",     "Customer & KYC"),
    # TRANSACTION_LEDGER
    "ACCOUNT_ID":        ("Account ID",               "Unique identifier for the financial account associated with a transaction",    "Payment Processing",   "Transactions"),
    "TRANSACTION_TYPE":  ("Transaction Type",         "Classification of the transaction: DEBIT, CREDIT, TRANSFER, FEE",             "Payment Processing",   "Transactions"),
    "MERCHANT_NAME":     ("Merchant Name",            "Name of the merchant or payee for the transaction",                           "Payment Processing",   "Transactions"),
    "MERCHANT_CATEGORY": ("Merchant Category",        "Merchant category classification used for spend analytics and reporting",      "Payment Processing",   "Transactions"),
    "BALANCE_AFTER":     ("Balance After Transaction","Account balance remaining immediately after a transaction is applied",         "Payment Processing",   "Transactions"),
    "CHANNEL":           ("Transaction Channel",      "Channel through which the transaction was initiated: ONLINE, BRANCH, ATM, MOBILE", "Payment Processing","Transactions"),
    "REFERENCE_NUMBER":  ("Reference Number",         "External reference or confirmation number assigned to the transaction",        "Payment Processing",   "Transactions"),
    # GL_ENTRY_REGISTER
    "ENTRY_ID":          ("GL Entry ID",              "Unique identifier for a general ledger journal entry",                        "Accounting Entries",   "General Ledger"),
    "JOURNAL_ID":        ("Journal ID",               "Identifier for the journal batch containing one or more GL entries",          "Accounting Entries",   "General Ledger"),
    "ENTRY_DATE":        ("Entry Date",               "Date on which the accounting entry was recorded",                             "Accounting Entries",   "General Ledger"),
    "ACCOUNT_NAME":      ("Account Name",             "Descriptive name of the general ledger account from the chart of accounts",   "Accounting Entries",   "General Ledger"),
    "NET_AMOUNT":        ("Net Amount",               "Net monetary amount calculated as debit minus credit for a GL entry",         "Accounting Entries",   "General Ledger"),
    "COST_CENTER":       ("Cost Center",              "Business unit cost center code used for financial reporting and allocation",   "Financial Close",      "General Ledger"),
    "LEGAL_ENTITY":      ("Legal Entity Code",        "Legal entity associated with the journal entry for consolidation reporting",  "Financial Close",      "General Ledger"),
    "FISCAL_YEAR":       ("Fiscal Year",              "Fiscal year in which the journal entry is recorded",                         "Financial Close",      "General Ledger"),
    "CREATED_BY":        ("Created By",               "User who created the journal entry in the financial system",                  "Accounting Entries",   "General Ledger"),
    "APPROVED_BY":       ("Approved By",              "User who approved the journal entry before posting to the ledger",            "Accounting Entries",   "General Ledger"),
    # RISK_EXPOSURE_DAILY
    "RISK_ID":           ("Risk Assessment ID",       "Unique identifier for a risk exposure assessment record",                     "Credit Risk",          "Risk & Regulatory"),
    "ASSESSMENT_DATE":   ("Assessment Date",          "Date on which the risk exposure assessment was performed",                    "Credit Risk",          "Risk & Regulatory"),
    "RISK_CATEGORY":     ("Risk Category",            "Type of risk: CREDIT, MARKET, OPERATIONAL, LIQUIDITY, REGULATORY",           "Credit Risk",          "Risk & Regulatory"),
    "EXPOSURE_AMOUNT":   ("Exposure Amount",          "Total monetary exposure amount in USD for the risk assessment",               "Credit Risk",          "Risk & Regulatory"),
    "EXPECTED_LOSS":     ("Expected Loss",            "Calculated expected loss: exposure × probability of default × loss given default", "Credit Risk",     "Risk & Regulatory"),
    "RISK_RATING":       ("Risk Rating",              "Internal credit risk rating assigned to the obligor: AAA through CCC",        "Credit Risk",          "Risk & Regulatory"),
    "COLLATERAL_VALUE":  ("Collateral Value",         "Estimated market value of collateral securing the exposure",                  "Credit Risk",          "Risk & Regulatory"),
    "REGULATORY_CAPITAL":("Regulatory Capital",       "Required regulatory capital allocation under Basel III rules",                "Regulatory Reporting", "Risk & Regulatory"),
    "BASEL_CLASS":       ("Basel Asset Class",        "Basel III asset class classification for capital adequacy calculation",       "Regulatory Reporting", "Risk & Regulatory"),
}

# ── Args ──────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--apply",    action="store_true", help="Phase 2: link Tab 1 approvals + generate import file")
parser.add_argument("--link-new", action="store_true", help="Phase 3: link Tab 2 approvals after import")
parser.add_argument("--company",  default="", help="Customer name for output filename")
parser.add_argument("--prefix",   default="", help="Asset prefix, e.g. ACM → ACMBT-GAP-N")
parser.add_argument("--out",      default="", help="Override output path")
parser.add_argument("--email",    default="", help="Governance owner email for new terms")
parser.add_argument("--yes",      action="store_true", help="Auto-approve all rows — skip workbook review")
args = parser.parse_args()

if args.out:
    OUT_PATH = Path(args.out)
elif args.company:
    slug = args.company.strip().replace(" ", "_")
    OUT_PATH = Path.home() / "Downloads" / f"CDGC_Gap_Review_{slug}.xlsx"
else:
    OUT_PATH = Path.home() / "Downloads" / "CDGC_Gap_Review.xlsx"

IMPORT_PATH = OUT_PATH.parent / "11_Business_Term_Gap.xlsx"

# ── Auth ──────────────────────────────────────────────────────────────────────
username = input("IDMC Username: ")
password = getpass.getpass("IDMC Password: ")

resp = requests.post(f"{LOGIN_URL}/identity-service/api/v1/Login",
    json={"username": username, "password": password}, timeout=30)
resp.raise_for_status()
data = resp.json()
session_id, org_id = data["sessionId"], data["orgId"]
resp = requests.get(
    f"{LOGIN_URL}/identity-service/api/v1/jwt/Token?client_id=idmc_api&nonce=1234",
    headers={"IDS-SESSION-ID": session_id},
    cookies={"USER_SESSION": session_id}, timeout=30)
resp.raise_for_status()
jwt = (resp.json().get("token") or resp.json().get("jwt_token")
       or resp.json().get("access_token"))
print("✓ Authenticated\n")

H  = {"Authorization": f"Bearer {jwt}", "X-INFA-ORG-ID": org_id,
      "Content-Type": "application/json", "Accept": "application/json"}

# ── API helpers ───────────────────────────────────────────────────────────────
def fetch_all(class_type):
    results, offset = [], 0
    while True:
        body = {"from": offset, "size": 100,
                "filterSpec": [{"type": "simple", "attribute": "core.classType",
                                 "values": [class_type]}]}
        r = requests.post(
            f"{ORG_URL}/data360/search/v1/assets?knowledgeQuery=*&segments=summary",
            headers=H, json=body, timeout=30)
        r.raise_for_status()
        hits = r.json().get("hits", [])
        results.extend(hits)
        if len(hits) < 100:
            break
        offset += 100
        time.sleep(0.2)
    return results


def patch_glossary(col_int_id, bt_ext_id):
    url  = f"{ORG_URL}/data360/content/v1/assets/{col_int_id}?scheme=internal"
    body = [{"operation": "add", "segment": "glossary",
             "items": [{"core.externalId": bt_ext_id}]}]
    r = requests.patch(url, headers=H, json=body, timeout=90)
    return r.status_code, r.text


def find_term_by_name(name):
    """Find a customer-authored Business Term by exact name. Skips system BT- terms."""
    r = requests.post(
        f"{ORG_URL}/data360/search/v1/assets?knowledgeQuery={requests.utils.quote(name)}&segments=summary",
        headers=H,
        json={"from": 0, "size": 10,
              "filterSpec": [{"type": "simple", "attribute": "core.classType",
                              "values": [TERM_CLASS]}]}, timeout=30)
    if r.status_code != 200:
        return None, None
    for hit in r.json().get("hits", []):
        s = hit.get("summary") or {}
        hit_name = s.get("core.name", "")
        ext_id   = hit.get("core.externalId") or s.get("core.externalId", "")
        if hit_name.lower() == name.lower() and not ext_id.startswith("BT-"):
            return ext_id, hit_name
    return None, None


def find_subdomains():
    """Fetch all subdomains: name → ext_id."""
    hits = fetch_all(SUBDOM_CLASS)
    result = {}
    for h in hits:
        s = h.get("summary") or {}
        ext_id = h.get("core.externalId") or s.get("core.externalId", "")
        name   = s.get("core.name", "")
        if ext_id and name:
            result[name] = ext_id
    return result

# ── Fuzzy matcher ─────────────────────────────────────────────────────────────
NOISE = {"of","the","a","an","and","or","for","in","on","at","to","is","as",
         "by","be","its","per","from","with","after","before"}

def tokens(s):
    return set(w.lower() for w in re.split(r'[_\s]+', s) if w)

def score(col_name, term_name, term_desc=""):
    ct = tokens(col_name)
    tt = tokens(term_name)
    if not ct or not tt:
        return 0
    if ct == tt:
        return 95
    if ct.issubset(tt):
        return 85
    if tt.issubset(ct):
        return 80
    overlap = ct & tt
    meaningful = overlap - NOISE
    if not meaningful:
        return 0
    jaccard = len(overlap) / len(ct | tt)
    desc_boost = 0
    if term_desc:
        dt = tokens(term_desc)
        if len(ct & dt - NOISE) >= len(ct - NOISE):
            desc_boost = 15
        elif ct & dt - NOISE:
            desc_boost = 8
    return min(100, int(jaccard * 75) + desc_boost)

def best_match(col_name, terms):
    best = None
    for t in terms:
        s = score(col_name, t["name"], t.get("description", ""))
        if s >= CONFIDENCE_THRESHOLD:
            if best is None or s > best[2]:
                best = (t["ext_id"], t["name"], s)
    return best

# ── Excel styles ──────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1B2A4A")
HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
YES_FILL  = PatternFill("solid", fgColor="E8F5E9")
NEW_FILL  = PatternFill("solid", fgColor="E3F2FD")
DONE_FILL = PatternFill("solid", fgColor="F1F8E9")
WARN_FILL = PatternFill("solid", fgColor="FFF8E1")
YES_FONT  = Font(color="2E7D32", bold=True)
BLUE_FONT = Font(color="1565C0", bold=True)
THIN      = Side(style="thin", color="DDE1E8")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, cols):
    ws.row_dimensions[1].height = 22
    for i, (label, width) in enumerate(cols, 1):
        c = ws.cell(1, i, label)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

def write_row(ws, rn, values, fill=None, approve_font=None):
    for i, val in enumerate(values, 1):
        c = ws.cell(rn, i, val)
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=(i >= len(values) - 1))
        if fill:
            c.fill = fill
    if approve_font:
        ws.cell(rn, 1).font = approve_font
    ws.row_dimensions[rn].height = 18

def inst_row(ws, text, ncols):
    ws.insert_rows(2)
    c = ws.cell(2, 1, text)
    c.font = Font(color="1565C0", italic=True, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — Link new terms after import (--link-new)
# ══════════════════════════════════════════════════════════════════════════════
if args.link_new:
    if not OUT_PATH.exists():
        print(f"ERROR: {OUT_PATH} not found — run analyze first")
        raise SystemExit(1)

    wb  = load_workbook(OUT_PATH)
    if "Suggested New Terms" not in wb.sheetnames:
        print("ERROR: 'Suggested New Terms' tab not found in workbook")
        raise SystemExit(1)

    ws2  = wb["Suggested New Terms"]
    hdrs = [c.value for c in ws2[1]]
    idx  = {h: i for i, h in enumerate(hdrs)}
    ok = fail = skip = 0

    print("Phase 3 — Linking new terms after import...")
    print("(Terms must already exist in CDGC from 11_Business_Term_Gap.xlsx import)\n")

    for row in ws2.iter_rows(min_row=3, values_only=True):
        approved = args.yes or str(row[idx.get("APPROVE", 0)] or "").strip().upper() == "YES"
        if not approved:
            skip += 1
            continue
        col_id    = row[idx["Column Internal ID"]]
        col_name  = row[idx["Column Name"]]
        table     = row[idx["Table"]]
        term_name = row[idx["Business Term Name"]]

        ext_id, found_name = find_term_by_name(term_name)
        if not ext_id:
            print(f"  ✗ {table}.{col_name} — term {term_name!r} not found in CDGC (imported yet?)")
            fail += 1
            time.sleep(0.2)
            continue

        status, body = patch_glossary(col_id, ext_id)
        already = status == 500 and ext_id in body and "RELATIONSHIP_ALREADY_EXISTS" in body
        if status in (200, 204) or already:
            tag = " (already linked)" if already else ""
            print(f"  ✓ {table}.{col_name} → {found_name} ({ext_id}){tag}")
            ok += 1
        else:
            print(f"  ✗ {table}.{col_name} → {ext_id}  [{status}] {body[:100]}")
            fail += 1
        time.sleep(0.3)

    print(f"\n{'─'*50}")
    print(f"Done.  {ok} linked,  {fail} failed,  {skip} skipped (not approved).")
    raise SystemExit(0)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — Link existing terms + generate import file (--apply)
# ══════════════════════════════════════════════════════════════════════════════
if args.apply:
    if not OUT_PATH.exists():
        print(f"ERROR: {OUT_PATH} not found — run analyze first")
        raise SystemExit(1)

    wb  = load_workbook(OUT_PATH)
    ok = fail = skip = 0

    # ── Tab 1: PATCH columns to existing terms ────────────────────────────────
    if "Suggested Links" in wb.sheetnames:
        ws1  = wb["Suggested Links"]
        hdrs = [c.value for c in ws1[1]]
        idx  = {h: i for i, h in enumerate(hdrs)}
        print("Phase 2 — Linking Tab 1 (existing terms)...\n")
        for row in ws1.iter_rows(min_row=3, values_only=True):
            approved = args.yes or str(row[idx.get("APPROVE", 0)] or "").strip().upper() == "YES"
            if not approved:
                skip += 1
                continue
            col_id   = row[idx["Column Internal ID"]]
            term_ext = row[idx["Term ext_id"]]
            col_name = row[idx["Column Name"]]
            table    = row[idx["Table"]]
            status, body = patch_glossary(col_id, term_ext)
            already = status == 500 and term_ext in body and "RELATIONSHIP_ALREADY_EXISTS" in body
            if status in (200, 204) or already:
                tag = " (already)" if already else ""
                print(f"  ✓ {table}.{col_name} → {term_ext}{tag}")
                ok += 1
            else:
                print(f"  ✗ {table}.{col_name}  [{status}] {body[:80]}")
                fail += 1
            time.sleep(0.3)

    print(f"\nTab 1 done: {ok} linked, {fail} failed, {skip} skipped\n")

    # ── Tab 2: generate 11_Business_Term_Gap.xlsx for CDGC bulk import ────────
    if "Suggested New Terms" in wb.sheetnames:
        ws2  = wb["Suggested New Terms"]
        hdrs = [c.value for c in ws2[1]]
        idx  = {h: i for i, h in enumerate(hdrs)}

        approved_new = []
        for row in ws2.iter_rows(min_row=3, values_only=True):
            approved = args.yes or str(row[idx.get("APPROVE", 0)] or "").strip().upper() == "YES"
            if not approved:
                continue
            approved_new.append({
                "col_id":    row[idx["Column Internal ID"]],
                "col_name":  row[idx["Column Name"]],
                "table":     row[idx["Table"]],
                "term_name": row[idx["Business Term Name"]],
                "term_desc": row[idx["Description"]],
                "subdomain": row[idx["Subdomain"]],
                "domain":    row[idx["Domain"]],
            })

        if approved_new:
            # Prompt for governance owner email if not supplied via --email
            gov_email = args.email.strip()
            if not gov_email:
                gov_email = input("Governance owner email (for new term imports): ").strip()

            # Fetch live subdomains to build ref_id lookup
            print("\nFetching subdomains from CDGC...")
            subdom_lookup = find_subdomains()
            print(f"  {len(subdom_lookup)} subdomains found")

            imp_wb = Workbook()
            ws_imp = imp_wb.active
            ws_imp.title = "Business Term"

            IMP_COLS = [
                ("Reference ID", 16), ("Name", 30), ("Description", 55),
                ("Alias Names", 20), ("Business Logic", 20), ("Critical Data Element", 22),
                ("Examples", 20), ("Format Type", 14), ("Format Description", 20),
                ("Lifecycle", 14), ("Security Level", 16), ("Classifications", 20),
                ("Reference Data", 20), ("Operation", 12),
                ("Parent: Subdomain", 35), ("Parent: Business Term", 20),
                ("Parent: Metric", 20), ("Parent: Domain", 20),
                ("Stakeholder: Governance Owner", 35),
                ("Stakeholder: Governance Administrator", 35),
            ]
            style_header(ws_imp, IMP_COLS)

            # Deduplicate by term name — same column name in multiple tables = one term
            seen_terms = {}
            for row in approved_new:
                name = row["term_name"]
                if name not in seen_terms:
                    seen_terms[name] = row

            for i, (name, row) in enumerate(seen_terms.items(), 1):
                subdom_name = row["subdomain"] or ""
                subdom_ext  = subdom_lookup.get(subdom_name, "")
                parent_ref  = f"{subdom_name} | {subdom_ext}" if subdom_ext else subdom_name

                write_row(ws_imp, ws_imp.max_row + 1, [
                    f"{args.prefix}BT-GAP-{i:03d}",   # Reference ID placeholder
                    name,                    # Name
                    row["term_desc"],        # Description
                    "",                      # Alias Names
                    "",                      # Business Logic
                    "false",                 # Critical Data Element
                    "",                      # Examples
                    "Text",                  # Format Type
                    "",                      # Format Description
                    "Published",             # Lifecycle
                    "Internal",              # Security Level
                    "",                      # Classifications
                    "",                      # Reference Data
                    "Create",                # Operation
                    parent_ref,              # Parent: Subdomain
                    "",                      # Parent: Business Term
                    "",                      # Parent: Metric
                    "",                      # Parent: Domain
                    gov_email,               # Stakeholder: Governance Owner
                    gov_email,               # Stakeholder: Governance Administrator
                ], fill=NEW_FILL)

            imp_wb.save(IMPORT_PATH)
            print(f"\n✓ Generated: {IMPORT_PATH}")
            print(f"  {len(seen_terms)} new Business Terms ready for import")
            subprocess.Popen(["open", str(IMPORT_PATH)])

            print(f"""
{'─'*60}
Next steps:

  1. Import {IMPORT_PATH.name} into CDGC:
     CDGC UI → Gear icon → Import → Upload → select the file → Import
     Wait for status: COMPLETED

  2. After import completes, run Phase 3 to link columns:
     python3 cdgc_gap_analyzer.py --link-new
{'─'*60}""")
        else:
            print("No approved new terms in Tab 2 — nothing to generate.")

    raise SystemExit(0)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — Analyze (default)
# ══════════════════════════════════════════════════════════════════════════════
print("Fetching scanned columns...")
col_hits = fetch_all(COLUMN_CLASS)
print(f"  {len(col_hits)} columns")

print("Fetching Business Terms...")
term_hits = fetch_all(TERM_CLASS)
terms = []
for h in term_hits:
    s      = h.get("summary") or {}
    ext_id = h.get("core.externalId") or s.get("core.externalId", "")
    name   = s.get("core.name", "")
    desc   = s.get("core.description", "")
    # Skip system-generated Glossary Association terms
    if ext_id and name and not ext_id.startswith("BT-"):
        terms.append({"ext_id": ext_id, "name": name, "description": desc})
print(f"  {len(terms)} terms (customer-authored)")

print("Fetching Subdomains...")
subdom_lookup = find_subdomains()
print(f"  {len(subdom_lookup)} subdomains\n")

# Build column records
columns = []
for h in col_hits:
    s      = h.get("summary") or {}
    name   = s.get("core.name", "")
    loc    = s.get("core.location", "")
    parts  = loc.rstrip("/").split("/")
    table  = parts[-2] if len(parts) >= 2 else ""
    int_id = h.get("core.identity", "")
    columns.append({"name": name, "table": table, "id": int_id})

# Categorize
linked_tab, suggest_link, suggest_new = [], [], []

for col in sorted(columns, key=lambda c: (c["table"], c["name"])):
    upper = col["name"].upper()
    if upper in ALREADY_LINKED_MAP:
        bt_ext, bt_name = ALREADY_LINKED_MAP[upper]
        linked_tab.append({**col, "term_ext": bt_ext, "term_name": bt_name})
    else:
        match = best_match(col["name"], terms)
        if match:
            suggest_link.append({**col, "term_ext": match[0],
                                  "term_name": match[1], "confidence": match[2]})
        else:
            draft = NEW_TERM_DRAFTS.get(upper)
            if draft:
                term_name, term_desc, subdomain, domain = draft
            else:
                term_name = col["name"].replace("_", " ").title()
                term_desc = f"[Edit: description for {term_name}]"
                subdomain = ""
                domain    = "Customer & KYC"
            suggest_new.append({**col,
                                 "draft_name": term_name,
                                 "draft_desc": term_desc,
                                 "subdomain":  subdomain,
                                 "domain":     domain})

print(f"Results:")
print(f"  Already governed : {len(linked_tab)}")
print(f"  Suggested links  : {len(suggest_link)}")
print(f"  Suggested new    : {len(suggest_new)}")
print()

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ── Tab 1: Suggested Links ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Suggested Links"
ws1.freeze_panes = "A3"
COLS1 = [("APPROVE",14),("Table",22),("Column Name",22),("Suggested Business Term",28),
         ("Confidence",12),("Term Description",42),("Term ext_id",14),("Column Internal ID",38)]
style_header(ws1, COLS1)
inst_row(ws1, "→ Set APPROVE to YES, then run: python3 cdgc_gap_analyzer.py --apply", len(COLS1))

for row in suggest_link:
    desc = next((t["description"] for t in terms if t["ext_id"] == row["term_ext"]), "")
    fill = YES_FILL if row["confidence"] >= 85 else WARN_FILL
    write_row(ws1, ws1.max_row+1,
              ["", row["table"], row["name"], row["term_name"],
               f"{row['confidence']}%", desc, row["term_ext"], row["id"]],
              fill=fill, approve_font=YES_FONT)

if not suggest_link:
    ws1.cell(ws1.max_row+1, 1, "No suggested links — all ungoverned columns need new terms (see Tab 2)")

# ── Tab 2: Suggested New Terms ────────────────────────────────────────────────
ws2 = wb.create_sheet("Suggested New Terms")
ws2.freeze_panes = "A3"
COLS2 = [("APPROVE",14),("Table",22),("Column Name",22),("Business Term Name",28),
         ("Description",52),("Subdomain",28),("Domain",22),("Column Internal ID",38)]
style_header(ws2, COLS2)
inst_row(ws2, "→ Review/edit, set APPROVE=YES, then run: python3 cdgc_gap_analyzer.py --apply", len(COLS2))

for row in suggest_new:
    write_row(ws2, ws2.max_row+1,
              ["", row["table"], row["name"], row["draft_name"],
               row["draft_desc"], row["subdomain"], row["domain"], row["id"]],
              fill=NEW_FILL, approve_font=BLUE_FONT)

# ── Tab 3: Already Governed ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Already Governed")
ws3.freeze_panes = "A2"
COLS3 = [("Table",22),("Column Name",22),("Business Term",28),("Term ext_id",14)]
style_header(ws3, COLS3)
for row in linked_tab:
    write_row(ws3, ws3.max_row+1,
              [row["table"], row["name"], row["term_name"], row["term_ext"]],
              fill=DONE_FILL)

wb.save(OUT_PATH)
print(f"✓ Written: {OUT_PATH}")
subprocess.Popen(["open", str(OUT_PATH)])
print()
print("Next steps:")
print(f"  1. Review {OUT_PATH.name} (opening now)")
print(f"  2. Tab 1 — set APPROVE=YES on suggested links to existing terms")
print(f"  3. Tab 2 — review AI-drafted definitions, set APPROVE=YES, edit Subdomain/Domain")
print(f"  4. Run Phase 2:  python3 cdgc_gap_analyzer.py --apply --company \"{args.company or 'Company Name'}\"")
print(f"     This links Tab 1 columns and generates 11_Business_Term_Gap.xlsx for import")
print(f"  5. Import 11_Business_Term_Gap.xlsx via CDGC UI")
print(f"  6. Run Phase 3:  python3 cdgc_gap_analyzer.py --link-new")
print(f"     This links Tab 2 columns to the newly imported terms")
