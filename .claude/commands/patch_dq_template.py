#!/usr/bin/env python3
"""
patch_dq_template.py

Applies ICDQ rule IDs to 13_DQ_Rule_Template.xlsx using fuzzy name matching.

Sets:
  - Technical Rule Reference = ICDQ artifact ID
  - Measuring Method = InformaticaCloudDataQuality

Rows with no ICDQ match are left unchanged (TechnicalScript).

USAGE
-----
  python3 patch_dq_template.py

  Prompts for:
    - Path to 13_DQ_Rule_Template.xlsx
    - Path to icdq_rules.csv (output of fetch_icdq_rule_ids.py --csv)

  Prints a proposed match table and requires you to type CONFIRM before writing.
"""
import csv
import re
import shutil
import openpyxl
from pathlib import Path

# ── File paths ────────────────────────────────────────────────────────────────
template_path = input("Path to 13_DQ_Rule_Template.xlsx: ").strip()
csv_path      = input("Path to icdq_rules.csv (from fetch_icdq_rule_ids.py): ").strip()
TEMPLATE = Path(template_path).expanduser()
CSV      = Path(csv_path).expanduser()
OUTPUT   = TEMPLATE.parent / "13_DQ_Rule_Template_PATCHED.xlsx"

# ── Fuzzy normalization ───────────────────────────────────────────────────────
def normalize(s):
    """Lowercase, remove underscores/hyphens/spaces for fuzzy comparison."""
    return re.sub(r'[\s_\-]+', '', s.lower())

# ── Load ICDQ rules ───────────────────────────────────────────────────────────
icdq = {}  # name -> id
with open(CSV) as f:
    for row in csv.DictReader(f):
        icdq[row["Name"]] = row["ID"]

# ── Load template names ───────────────────────────────────────────────────────
wb_read = openpyxl.load_workbook(TEMPLATE, read_only=True)
ws_read = wb_read.active
headers = [cell.value for cell in next(ws_read.iter_rows(min_row=1, max_row=1))]
col_ref_idx  = headers.index("Reference ID")
col_name_idx = headers.index("Name")

template_rows = []
for row in ws_read.iter_rows(min_row=2, values_only=True):
    ref_id = row[col_ref_idx]
    name   = row[col_name_idx]
    if ref_id and name:
        template_rows.append((str(ref_id).strip(), str(name).strip()))
wb_read.close()

# ── Build fuzzy matches ───────────────────────────────────────────────────────
norm_icdq = {normalize(k): (k, v) for k, v in icdq.items()}
proposed = {}  # ref_id -> (icdq_name, rule_id) or None

print(f"\n{'─'*100}")
print(f"  {'Ref ID':<12} {'Template Name':<40} {'Best ICDQ Match':<40} {'Match?'}")
print(f"{'─'*100}")

for ref_id, tname in template_rows:
    norm_t = normalize(tname)
    # Try exact match first, then substring match
    match = norm_icdq.get(norm_t)
    if not match:
        # Find ICDQ rule whose normalized name contains the normalized template name or vice versa
        candidates = [(k, v) for k, v in norm_icdq.items()
                      if norm_t in k or k in norm_t]
        if candidates:
            # Pick shortest (most specific) match
            best_key = min(candidates, key=lambda x: len(x[0]))
            match = best_key[1]
    if match:
        icdq_name, rule_id = match
        proposed[ref_id] = (icdq_name, rule_id)
        print(f"  {ref_id:<12} {tname:<40} {icdq_name:<40} ✓")
    else:
        proposed[ref_id] = None
        print(f"  {ref_id:<12} {tname:<40} {'(no match — TechnicalScript)':<40}")

print(f"{'─'*100}")
matched_count = sum(1 for v in proposed.values() if v)
print(f"\n  Matched: {matched_count}  |  No match (TechnicalScript): {len(proposed) - matched_count}")
print(f"\nReview the matches above carefully.")
confirm = input("Type CONFIRM to apply, or anything else to abort: ").strip()
if confirm != "CONFIRM":
    raise SystemExit("Aborted.")

# ── Patch the workbook ────────────────────────────────────────────────────────
shutil.copy(TEMPLATE, OUTPUT)
wb = openpyxl.load_workbook(OUTPUT)
ws = wb.active

hdrs       = [cell.value for cell in ws[1]]
col_method = hdrs.index("Measuring Method") + 1
col_ref    = hdrs.index("Technical Rule Reference") + 1
col_op     = hdrs.index("Operation") + 1
col_out    = hdrs.index("Output Port Name") + 1

print(f"\n{'ID':<12} {'Template Name':<45} {'ICDQ Rule':<45} {'Ref ID'}")
print("─" * 130)

matched = 0
skipped = []
for row in ws.iter_rows(min_row=2):
    ref_id = row[0].value
    name   = row[1].value
    if not ref_id:
        continue

    # All rows already exist in CDGC — force Update
    ws.cell(row=row[0].row, column=col_op).value = "Update"

    match = proposed.get(str(ref_id).strip())
    if match:
        icdq_name, rule_id = match
        ws.cell(row=row[0].row, column=col_method).value = "InformaticaCloudDataQuality"
        ws.cell(row=row[0].row, column=col_ref).value    = rule_id
        ws.cell(row=row[0].row, column=col_out).value    = "Output"
        print(f"  {ref_id:<10} {name:<45} {icdq_name:<45} {rule_id}")
        matched += 1
    else:
        print(f"  {ref_id:<10} {name:<45} (no ICDQ match — left as TechnicalScript)")
        skipped.append(str(ref_id).strip())

wb.save(OUTPUT)
print(f"\n{'─' * 130}")
print(f"✓ Saved: {OUTPUT}")
print(f"  Matched: {matched}   Left as TechnicalScript: {len(skipped)}")
if skipped:
    print(f"  Skipped: {', '.join(skipped)}")
